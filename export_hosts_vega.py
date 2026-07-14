import urllib.request
import json
import sys
import os

# Auto-instala a biblioteca openpyxl para gerar planilhas reais com abas e cores
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Biblioteca 'openpyxl' nao encontrada. Instalando automaticamente...")
    import subprocess
    try:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        print("Biblioteca 'openpyxl' instalada com sucesso!\n")
    except Exception as e:
        print(f"Erro ao instalar 'openpyxl': {e}")
        print("Por favor, execute manualmente no terminal: pip install openpyxl")
        sys.exit(1)

# CONFIGURAÇÕES DO SEU ZABBIX - COLOQUE O SEU TOKEN
TOKEN = "SEU_TOKEN_AQUI"
BASE_URL = "http://10.10.1.26"

URLS_PARA_TENTAR = [
    f"{BASE_URL}/api_jsonrpc.php",
    f"{BASE_URL}/zabbix/api_jsonrpc.php"
]

print("Iniciando conexao com a API do Zabbix...")

payload_hosts = {
    "jsonrpc": "2.0",
    "method": "host.get",
    "params": {
        "output": [
            "hostid", "host", "name", "status", "proxy_hostid", "description"
        ],
        "selectInterfaces": ["ip", "type", "port"],
        "selectHostGroups": ["name"],
        "selectParentTemplates": ["name"],
        "selectInventory": ["os", "hardware", "software", "serialno_a"]
    },
    "id": 1
}

data_hosts = json.dumps(payload_hosts).encode('utf-8')
headers = {
    "Content-Type": "application/json",
    "Authorization": f"Bearer {TOKEN}"
}

sucesso = False
hosts = None
url_conectada = ""

for url in URLS_PARA_TENTAR:
    print(f"Tentando conectar em: {url}...")
    req = urllib.request.Request(url, data=data_hosts, headers=headers)
    try:
        with urllib.request.urlopen(req, timeout=15) as response:
            res_data = response.read().decode('utf-8')
            res_json = json.loads(res_data)
            if 'result' in res_json:
                hosts = res_json['result']
                sucesso = True
                url_conectada = url
                print(f"[OK] Conectado com sucesso em: {url}")
                break
    except Exception as e:
        print(f"[FALHA REDE] {url}: {e}")

if not sucesso:
    print("[ERRO CRITICO] Nao foi possivel conectar à API do Zabbix.")
    sys.exit(1)

host_ids = [h['hostid'] for h in hosts]

proxies_map = {}
payload_proxies = {
    "jsonrpc": "2.0",
    "method": "proxy.get",
    "params": {"output": ["proxyid", "name"]},
    "id": 2
}
try:
    data_proxies = json.dumps(payload_proxies).encode('utf-8')
    req_proxies = urllib.request.Request(url_conectada, data=data_proxies, headers=headers)
    with urllib.request.urlopen(req_proxies, timeout=10) as response:
        res_proxies = json.loads(response.read().decode('utf-8'))
        if 'result' in res_proxies:
            for p in res_proxies['result']:
                proxies_map[str(p['proxyid'])] = p['name']
except Exception as e:
    print(f"[AVISO] Falha ao listar proxies: {e}")

triggers_map = {}
payload_triggers = {
    "jsonrpc": "2.0",
    "method": "trigger.get",
    "params": {
        "output": ["triggerid", "description", "priority"],
        "selectHosts": ["hostid"],
        "monitored": True,
        "only_true": True,
        "filter": {"value": 1}
    },
    "id": 3
}
try:
    data_triggers = json.dumps(payload_triggers).encode('utf-8')
    req_triggers = urllib.request.Request(url_conectada, data=data_triggers, headers=headers)
    with urllib.request.urlopen(req_triggers, timeout=15) as response:
        res_triggers = json.loads(response.read().decode('utf-8'))
        if 'result' in res_triggers:
            priorities = {
                "0": "Classif.", "1": "Info", "2": "Atencao", 
                "3": "Media", "4": "Alta", "5": "Desastre"
            }
            for t in res_triggers['result']:
                desc = t['description']
                priority_label = priorities.get(str(t['priority']), "Info")
                for h in t.get('hosts', []):
                    hid = h['hostid']
                    if hid not in triggers_map:
                        triggers_map[hid] = []
                    triggers_map[hid].append(f"[{priority_label.upper()}] {desc}")
except Exception as e:
    print(f"[AVISO] Falha ao listar alertas ativos: {e}")

payload_items = {
    "jsonrpc": "2.0",
    "method": "item.get",
    "params": {
        "hostids": host_ids,
        "output": ["hostid", "name", "key_", "lastvalue", "units"],
        "search": {
            "key_": [
                "cpu.util", "cpu.usage", "hrProcessorLoad", 
                "memory.util", "memory.size[pused]", "memory.used", 
                "pused",
                "system.sw.os", "system.descr", "system.uname",
                "system.hw.model", "system.hw.device",
                "system.uptime", "sysUpTimeInstance", "hrSystemUptime"
            ]
        },
        "searchByAny": True,
        "monitored": True
    },
    "id": 4
}
data_items = json.dumps(payload_items).encode('utf-8')
req_items = urllib.request.Request(url_conectada, data=data_items, headers=headers)

items = []
try:
    with urllib.request.urlopen(req_items, timeout=25) as response:
        res_json = json.loads(response.read().decode('utf-8'))
        if 'result' in res_json:
            items = res_json['result']
except Exception as e:
    print(f"[AVISO] Erro ao buscar itens: {e}")

host_items_map = {}
for item in items:
    hid = item['hostid']
    if hid not in host_items_map:
        host_items_map[hid] = []
    host_items_map[hid].append(item)

aba_hypervisors = []
aba_vms = []
aba_rede = []
aba_geral = []

for host in hosts:
    hostid = host.get('hostid', '')
    hostname = host.get('host', '')
    name = host.get('name', '')
    status = "Ativo" if host.get('status') == "0" else "Inativo"
    
    proxy_id = str(host.get('proxy_hostid', ''))
    proxy_name = proxies_map.get(proxy_id, "Direto (Zabbix Server)") if proxy_id != "0" and proxy_id != "" else "Direto (Zabbix Server)"
    
    ips = [iface.get('ip', '') for iface in host.get('interfaces', []) if iface.get('ip')]
    ip_str = ", ".join(list(set(ips))) if ips else ""
    
    types = []
    for iface in host.get('interfaces', []):
        iftype = iface.get('type')
        if iftype == '1': types.append("Agente Zabbix")
        elif iftype == '2': types.append("SNMP")
        elif iftype == '3': types.append("IPMI")
        elif iftype == '4': types.append("JMX")
    type_str = ", ".join(list(set(types))) if types else "API / Sem Agente"
    
    groups = [g.get('name', '') for g in host.get('hostgroups', [])]
    groups_str = ", ".join(groups)
    
    templates = [t.get('name', '') for t in host.get('parentTemplates', [])]
    templates_str = ", ".join(templates)
    
    alertas_list = triggers_map.get(hostid, [])
    active_alerts = " | ".join(alertas_list) if alertas_list else "Sem Incidentes"
    
    host_desc = host.get('description', '') or ""
    host_desc = host_desc.replace("\n", " ").replace("\r", " ").strip()

    inventory = host.get('inventory')
    os_inv = inventory.get('os', '') if isinstance(inventory, dict) else ""
    hw_inv = inventory.get('hardware', '') if isinstance(inventory, dict) else ""
    serial = inventory.get('serialno_a', '') if isinstance(inventory, dict) else ""
    
    host_items = host_items_map.get(hostid, [])
    cpu_val, mem_val, disks, os_fallback, hw_fallback, uptime_val = "N/A", "N/A", [], "", "", "N/A"
    
    for item in host_items:
        key = item.get('key_', '')
        val = item.get('lastvalue', '') or ""
        units = item.get('units', '') or ""
        item_name = item.get('name', '')
        
        if val == "": continue
        
        if key in ["system.sw.os", "system.descr", "system.uname"]:
            if not os_fallback or key == "system.sw.os":
                os_fallback = val
        if key in ["system.hw.model", "system.hw.device"]:
            hw_fallback = val
            
        try:
            val_float = float(val)
            val_str = f"{val_float:.2f}{units}" if units else f"{val_float:.2f}"
        except ValueError:
            val_str = f"{val}{units}" if units else f"{val}"
            
        if ("cpu.util" in key or "cpu.usage" in key or "hrProcessorLoad" in key) and not ("hrProcessorLoad[" in key):
            if cpu_val == "N/A" or "system.cpu.util" in key or "vm.cpu.util" in key:
                cpu_val = val_str
        elif "memory.util" in key or "memory.size[pused]" in key or "memory.used" in key:
            if mem_val == "N/A" or "util" in key or "pused" in key:
                mem_val = val_str
        elif "pused" in key:
            disk_label = key.split("[")[-1].split(",")[0].replace("]", "") if "[" in key else item_name
            disk_entry = f"{disk_label}: {val_str}"
            if disk_entry not in disks:
                disks.append(disk_entry)
        elif "system.uptime" in key or "sysUpTimeInstance" in key or "hrSystemUptime" in key:
            try:
                seconds = float(val)
                if "SNMP" in type_str and units != "s" and units != "uptime":
                    seconds = seconds / 100.0
                days = int(seconds // 86400)
                hours = int((seconds % 86400) // 3600)
                minutes = int((seconds % 3600) // 60)
                uptime_val = f"{days}d {hours}h {minutes}m" if days > 0 else (f"{hours}h {minutes}m" if hours > 0 else f"{minutes}m")
            except ValueError:
                uptime_val = val
                
    final_os = os_inv if os_inv else (os_fallback if os_fallback else "N/A")
    final_hw = hw_inv if hw_inv else (hw_fallback if hw_fallback else "N/A")
    final_os = final_os.replace("\n", " ").replace("\r", " ").strip()
    final_hw = final_hw.replace("\n", " ").replace("\r", " ").strip()
    disks_str = " | ".join(disks) if disks else "N/A"
    
    row_data = [
        hostid, hostname, name, ip_str, type_str, status, proxy_name,
        cpu_val, mem_val, disks_str, uptime_val, active_alerts, groups_str, templates_str,
        final_os, final_hw, serial, host_desc
    ]
    
    aba_geral.append(row_data)
    
    # Classificação
    if "Virtualization" in groups_str or "Hypervisors" in groups_str or "VMware Hypervisor" in templates_str or "ESXi" in templates_str or "idrac" in hostname.lower() or "idrac" in name.lower() or "dell" in templates_str.lower():
        aba_hypervisors.append(row_data)
    elif any(x in groups_str for x in ["Network", "Switches", "Routers", "Firewalls"]) or any(y in templates_str for y in ["SNMP", "FortiGate", "Cisco", "Switch"]):
        aba_rede.append(row_data)
    else:
        aba_vms.append(row_data)

wb = openpyxl.Workbook()
ws_dash = wb.active
ws_dash.title = "Resumo Geral"
ws_dash.views.sheetView[0].showGridLines = True

cor_header_fill = PatternFill(start_color="1E4A38", end_color="1E4A38", fill_type="solid")
cor_zebra_fill = PatternFill(start_color="F5F8F6", end_color="F5F8F6", fill_type="solid")
cor_white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
font_normal = Font(name="Calibri", size=11, color="000000")
font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")

thin_border_side = Side(border_style="thin", color="D3D3D3")
border_cell = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

# Layout de Dashboard
ws_dash.merge_cells("A1:D1")
ws_dash["A1"] = "DASHBOARD DE INVENTÁRIO - ZABBIX"
ws_dash["A1"].font = font_title
ws_dash["A1"].fill = cor_header_fill
ws_dash["A1"].alignment = align_center
ws_dash.row_dimensions[1].height = 40

ws_dash["A3"] = "Métrica de Infraestrutura"
ws_dash["B3"] = "Quantidade"
for col in ["A3", "B3"]:
    ws_dash[col].font = font_header
    ws_dash[col].fill = cor_header_fill
    ws_dash[col].alignment = align_center
    ws_dash[col].border = border_cell

kpis = [
    ("Total de Hosts Monitorados", len(hosts)),
    ("Hosts com Status Ativo", sum(1 for h in hosts if h['status'] == "0")),
    ("Hosts com Status Inativo", sum(1 for h in hosts if h['status'] != "0")),
    ("Total de Incidentes / Alertas Ativos", sum(len(x) for x in triggers_map.values()))
]
for idx, (label, val) in enumerate(kpis, start=4):
    ws_dash[f"A{idx}"] = label
    ws_dash[f"A{idx}"].font = font_normal
    ws_dash[f"A{idx}"].border = border_cell
    ws_dash[f"B{idx}"] = val
    ws_dash[f"B{idx}"].font = Font(name="Calibri", size=11, bold=True)
    ws_dash[f"B{idx}"].alignment = align_center
    ws_dash[f"B{idx}"].border = border_cell

colunas = [
    "ID do Host", "Nome Tecnico", "Nome Visivel", "Endereco IP", 
    "Modelo de Monitoramento", "Status (Ativo/Inativo)", "Proxy de Monitoramento",
    "CPU Ult. Valor", "Memoria Ult. Valor", "Discos (Uso %)", "Uptime", "Alertas Ativos no Ativo",
    "Grupos", "Templates Vinculados", "Sistema Operacional (OS)", "Hardware", "Numero de Serie", "Descricao"
]

def criar_aba(titulo, dados):
    ws = wb.create_sheet(title=titulo)
    ws.views.sheetView[0].showGridLines = True
    for col_idx, col_name in enumerate(colunas, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = font_header
        cell.fill = cor_header_fill
        cell.alignment = align_center
        cell.border = border_cell
    ws.row_dimensions[1].height = 28
    
    for row_idx, r_data in enumerate(dados, start=2):
        ws.row_dimensions[row_idx].height = 20
        fill_to_use = cor_zebra_fill if row_idx % 2 == 0 else cor_white_fill
        for col_idx, value in enumerate(r_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = font_normal
            cell.fill = fill_to_use
            cell.border = border_cell
            if col_idx in [1, 4, 6]: cell.alignment = align_center
            elif col_idx in [8, 9]: cell.alignment = align_right
            else: cell.alignment = align_left
            
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                val_len = len(str(cell.value))
                if col[0].column in [10, 12, 13, 14, 18]:
                    val_len = min(val_len, 30)
                max_len = max(max_len, val_len)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

criar_aba("Servidores Físicos (Host)", aba_hypervisors)
criar_aba("Máquinas Virtuais (VMs)", aba_vms)
criar_aba("Ativos de Rede (Switches)", aba_rede)
criar_aba("Todos os Ativos", aba_geral)

for col in ws_dash.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws_dash.column_dimensions[col_letter].width = max(max_len + 5, 15)

xlsx_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ativos_zabbix_vega.xlsx")
try:
    wb.save(xlsx_file)
    print(f"\n[SUCESSO] Planilha Excel (.xlsx) gerada: {xlsx_file}")
except Exception as e:
    print(f"\n[ERRO] Falha ao salvar a planilha: {e}")
    sys.exit(1)
